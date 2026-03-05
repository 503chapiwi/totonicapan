import pdfplumber
import openpyxl
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter
import unicodedata
import streamlit as st
import re
import io

# --- HELPER FUNCTIONS ---
def normalize_text(text):
    if not text: return ""
    nfd = unicodedata.normalize('NFD', str(text))
    return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn').lower()

def squish_text(text):
    """Aggressively removes ALL spaces, punctuation, hyphens, and hidden characters for a 100% reliable match."""
    if not text: return ""
    t = normalize_text(text)
    return re.sub(r'[^a-z0-9]', '', t)

def safe_float(val):
    if val is None: return 0.0
    s = str(val).strip()
    if not s or s == '-': return 0.0
    s = s.replace(',', '') 
    s = re.sub(r'[^\d\.\-]', '', s) 
    if s.count('.') > 1:
        parts = s.rsplit('.', 1)
        s = parts[0].replace('.', '') + '.' + parts[1]
    try: return float(s)
    except ValueError: return 0.0

def clean_currency(value):
    if not value: return 0.0
    raw = str(value).strip().replace(' ', '')
    raw = re.sub(r'[^\d\.,]', '', raw)
    if not raw: return 0.0
    
    if re.search(r',\d{1,2}$', raw):
        parts = raw.rsplit(',', 1)
        raw = parts[0].replace('.', '').replace(',', '') + '.' + parts[1]
    else:
        raw = raw.replace(',', '')
        
    if raw.count('.') > 1:
        parts = raw.rsplit('.', 1)
        raw = parts[0].replace('.', '') + '.' + parts[1]
        
    try: return float(raw)
    except ValueError: return 0.0

def extract_value_from_row(row_list, total_idx):
    if total_idx != -1 and len(row_list) > total_idx:
        val = clean_currency(row_list[total_idx])
        if val > 0: return val
    for item in reversed(row_list):
        val = clean_currency(item)
        if val > 0: return val
    return 0.0

def get_master_cell(ws, r_idx, c_idx):
    cell = ws.cell(row=r_idx, column=c_idx)
    if type(cell).__name__ == 'MergedCell':
        for m_range in ws.merged_cells.ranges:
            if cell.coordinate in m_range:
                return ws.cell(row=m_range.min_row, column=m_range.min_col)
    return cell

# --- TRUCO CSS PARA TRADUCIR LA INTERFAZ A ESPAÑOL ---
st.markdown("""
    <style>
        .stFileUploader > div > div > div > div > span:first-child { display: none; }
        .stFileUploader > div > div > div > div::before {
            content: "Arrastre y suelte los archivos aquí";
            display: block; font-weight: 600; margin-bottom: 5px;
        }
    </style>
""", unsafe_allow_html=True)

# --- WEB UI ---
st.title("🇬🇹 MAGA: Procesador de Facturas por la LAE")
uploaded_pdfs = st.file_uploader(label='1. Seleccione sus Facturas (PDFs)', type='pdf', accept_multiple_files=True)
uploaded_xlsx = st.file_uploader(label='2. Seleccione su Archivo de Excel', type='xlsx')

if st.button("INICIAR PROCESO") and uploaded_pdfs and uploaded_xlsx:
    try:
        input_buffer = io.BytesIO(uploaded_xlsx.read())
        wb = openpyxl.load_workbook(input_buffer)
        ws = wb.active 
        
        if "Extra Detalles" not in wb.sheetnames:
            ws_det = wb.create_sheet("Extra Detalles")
            ws_det.append(['Nombre Emisor', 'NIT Emisor', 'NIT Receptor', 'UUID', 'Municipio', 'Alerta % Abarrotes'])
        else:
            ws_det = wb["Extra Detalles"]

        # 1. Map Excel Columns dynamically
        col_map = {}
        for row in ws.iter_rows(min_row=1, max_row=15): 
            for cell in row:
                if type(cell).__name__ == 'MergedCell': continue
                if not cell.value: continue
                val = normalize_text(str(cell.value))
                
                if 'abarrotes' in val: col_map['abar'] = cell.column
                if 'agricultura' in val: col_map['agri'] = cell.column
                if 'escuela' in val or 'establecimiento' in val: col_map['escuelas'] = cell.column
                if 'proveedor' in val or 'productor' in val:
                    base_col, base_row, found_total = cell.column, cell.row, False
                    for r_offset in range(1, 4):
                        for c_offset in range(3):
                            sub_cell = ws.cell(row=base_row + r_offset, column=base_col + c_offset)
                            if sub_cell.value and 'total' in normalize_text(str(sub_cell.value)):
                                col_map['productores'] = sub_cell.column
                                found_total = True
                                break
                        if found_total: break
                    if 'productores' not in col_map: col_map['productores'] = base_col

        if 'abar' not in col_map or 'agri' not in col_map:
            st.error(f"No encontré las columnas base en el Excel.")
            st.stop()

        # 2. MASTER MUNICIPALITY DICTIONARY
        MUNICIPIOS = {
            1: {"nombre_oficial": "Totonicapán", "alias_pdf": ["totonicapan totonicapan", "totonicapan, totonicapan", "totonicapan"]},
            2: {"nombre_oficial": "San Cristóbal Totonicapán", "alias_pdf": ["san cristobal totonicapan", "san cristobal"]},
            3: {"nombre_oficial": "San Francisco El Alto", "alias_pdf": ["san francisco el alto", "san francisco"]},
            4: {"nombre_oficial": "San Andrés Xecul", "alias_pdf": ["san andres xecul", "san andres"]},
            5: {"nombre_oficial": "Momostenango", "alias_pdf": ["momostenango"]},
            6: {"nombre_oficial": "Santa María Chiquimula", "alias_pdf": ["santa maria chiquimula", "sta maria chiquimula", "santa maria", "sta maria"]},
            7: {"nombre_oficial": "Santa Lucía La Reforma", "alias_pdf": ["santa lucia la reforma", "sta lucia la reforma", "santa lucia", "sta lucia"]},
            8: {"nombre_oficial": "San Bartolo Aguas Calientes", "alias_pdf": ["san bartolo aguas calientes", "san bartolo"]}
        }
        
        search_list = []
        for m_id, data in MUNICIPIOS.items():
            for alias in data["alias_pdf"]:
                search_list.append((alias, m_id, data["nombre_oficial"]))
                
        # CORE FIX: Sorts the list so Totonicapán (ID 1) is ALWAYS evaluated last.
        # Within the other municipalities, sorts by length to catch specific names first.
        search_list.sort(key=lambda x: (x[1] == 1, -len(x[0])))

        EXCEL_MAPPINGS = {
            1: "totonicapán", 2: "san cristobal", 3: "san francisco", 4: "san andres",
            5: "momostenango", 6: "santa maria", 7: "santa lucia", 8: "san bartolo"
        }

        # 3. Map Excel Rows to Municipalities
        row_map = {}
        for row_ex in ws.iter_rows(min_row=5, max_row=150):
            row_text = " ".join([str(c.value) for c in row_ex if c.value and type(c).__name__ != 'MergedCell'])
            row_squished = squish_text(row_text)
            for m_id, search_key in EXCEL_MAPPINGS.items():
                if m_id in row_map: continue
                key_squished = squish_text(search_key)
                if key_squished in row_squished:
                    row_map[m_id] = row_ex[0].row

        batch_totals = {m_id: {'abar': 0.0, 'agri': 0.0, 'emisores': set(), 'receptores': set()} for m_id in MUNICIPIOS.keys()}
        new_count = 0
        progress_bar = st.progress(0)

        # 4. Process each PDF
        for i, pdf_file in enumerate(uploaded_pdfs):
            with pdfplumber.open(pdf_file) as pdf:
                text = "".join([p.extract_text() or "" for p in pdf.pages])
                tables = []
                for p in pdf.pages:
                    t = p.extract_table()
                    if t: tables.extend(t)

                uuid_m = re.search(r'\b[A-F0-9]{8}-[A-F0-9]{4}-[A-F0-9]{4}-[A-F0-9]{4}-[A-F0-9]{12}\b', text, re.I)
                uuid_val = uuid_m.group(0).upper() if uuid_m else pdf_file.name

                text_squished = squish_text(text)
                m_id, m_name = None, "N/A"
                
                # Check against our aggressively squished master list
                for alias, mun_id, official_name in search_list:
                    alias_squished = squish_text(alias)
                    if alias_squished in text_squished:
                        m_id = mun_id
                        m_name = official_name
                        break

                if m_id:
                    abar_sum, agri_sum = 0, 0
                    cultivados = ['tomate', 'pina', 'piña', 'banano', 'zanahoria', 'guisquil', 'cebolla', 'aguacate', 
                                  'miltomate', 'brocoli', 'melon', 'melón', 'ejote', 'maiz', 'maíz', 'jamaica', 
                                  'cebada', 'papaya', 'manzana', 'chile', 'apio', 'ajo', 'cilantro', 'tusa', 'sandia', 'sandía']
                    abarrotes = ['pollo', 'tostada', 'huevo', 'pan', 'queso', 'carne', 'res']
                    
                    total_col_idx = -1
                    for row_tbl in tables:
                        if not row_tbl: continue
                        for idx, cell in enumerate(row_tbl):
                            if cell and 'total' in normalize_text(str(cell)) and 'descuento' not in normalize_text(str(cell)):
                                total_col_idx = idx
                                break
                        if total_col_idx != -1: break

                    for row_tbl in tables:
                        if not row_tbl: continue
                        row_text = " ".join([normalize_text(str(x)) for x in row_tbl if x])
                        val = extract_value_from_row(row_tbl, total_col_idx)
                            
                        if any(x in row_text for x in cultivados): agri_sum += val
                        if any(x in row_text for x in abarrotes): abar_sum += val
                    
                    nit_e_match = re.search(r'Emisor:\s*([0-9Kk\-]+)', text, re.I)
                    nit_r_match = re.search(r'Receptor:\s*([0-9Kk\-]+)', text, re.I)
                    name_e_match = re.search(r'(?:Factura(?:\s*Pequeño\s*Contribuyente)?)\s*\n+(.*?)\n+Nit\s*Emisor', text, re.IGNORECASE | re.DOTALL)
                    
                    nit_e = nit_e_match.group(1).strip() if nit_e_match else "N/A"
                    nit_r = nit_r_match.group(1).strip() if nit_r_match else "N/A"
                    raw_name = re.sub(r'\s+', ' ', name_e_match.group(1).strip() if name_e_match else "N/A")
                    name_e = re.split(r'(?i)n[úu]mero\s*de\s*autorizaci[óo]n', raw_name)[0]
                    name_e = re.split(r'(?i)\bserie\b', name_e)[0].strip()

                    batch_totals[m_id]['abar'] += abar_sum
                    batch_totals[m_id]['agri'] += agri_sum
                    if nit_e != "N/A": batch_totals[m_id]['emisores'].add(nit_e)
                    if nit_r != "N/A": batch_totals[m_id]['receptores'].add(nit_r)

                    total_rec = abar_sum + agri_sum
                    perc_abar = (abar_sum / total_rec) if total_rec > 0 else 0
                    alert_status = "⚠️ ALERTA: >30%" if perc_abar > 0.30 else "OK"

                    ws_det.append([name_e, nit_e, nit_r, uuid_val, m_name, alert_status])
                    new_count += 1
                else:
                    st.warning(f"No se pudo identificar el municipio en la factura: {pdf_file.name}")

            progress_bar.progress((i + 1) / len(uploaded_pdfs))

        # 5. Write to Main Sheet securely
        for target_m_id, r_idx in row_map.items():
            data = batch_totals.get(target_m_id)
            if not data: continue

            if 'abar' in col_map and data['abar'] > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['abar'])
                target_cell.value = safe_float(target_cell.value) + data['abar']
            
            if 'agri' in col_map and data['agri'] > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['agri'])
                target_cell.value = safe_float(target_cell.value) + data['agri']

            if 'escuelas' in col_map and len(data['receptores']) > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['escuelas'])
                target_cell.value = int(safe_float(target_cell.value)) + len(data['receptores'])
            
            if 'productores' in col_map and len(data['emisores']) > 0:
                target_cell = get_master_cell(ws, r_idx, col_map['productores'])
                target_cell.value = int(safe_float(target_cell.value)) + len(data['emisores'])

        # 6. Format "Extra Detalles"
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for col in ws_det.columns:
            max_length = 0
            col_letter = get_column_letter(col[0].column) 
            for cell in col:
                cell.border = thin_border 
                try: max_length = max(max_length, len(str(cell.value)))
                except: pass
            ws_det.column_dimensions[col_letter].width = max_length + 2

        # 7. Final Export
        output = io.BytesIO()
        wb.save(output)
        
        st.success(f"¡Proceso completado! {new_count} facturas procesadas y agregadas al Excel con éxito.")
        output.seek(0)
        st.download_button("Descargar Reporte Final", data=output.getvalue(), 
                           file_name="Reporte_MAGA_Actualizado.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    except Exception as e:
        st.error(f"Error crítico detectado: {e}")
